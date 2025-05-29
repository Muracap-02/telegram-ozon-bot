
import os
import pandas as pd
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, MessageHandler, filters,
    ContextTypes, CommandHandler, CallbackQueryHandler, ConversationHandler
)
import tempfile
import logging
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import zipfile
from openpyxl.cell.cell import MergedCell

logging.basicConfig(format='[LOG] %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

TEMPLATE_FILENAME = "AllPackageEC_.xlsx"
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), TEMPLATE_FILENAME)

MODE_CHOICE = {}
PINFL_STEP, PINFL_SOURCE, PINFL_PINFL = range(3)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("▶️ Обработать на части (1000)", callback_data="chunk")],
        [InlineKeyboardButton("▶️ Обработать на части (500)", callback_data="chunk500")],
        [InlineKeyboardButton("▶️ Обработать на части (250)", callback_data="chunk250")],
        [InlineKeyboardButton("📄 Макрос Пасспорт", callback_data="passport")],
        [InlineKeyboardButton("🔄 Замена ПИНФЛ", callback_data="pinfl_replace")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите режим обработки:", reply_markup=reply_markup)

async def mode_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    MODE_CHOICE[user_id] = query.data

    if query.data == "pinfl_replace":
        await query.message.reply_text("Пожалуйста, загрузите файл реестра (source).")
        context.user_data['pinfl_step'] = PINFL_SOURCE
        return PINFL_STEP
    else:
        await query.message.reply_text("Отправьте Excel-файл для выбранной обработки.")
        return ConversationHandler.END

async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.message.from_user
    user_id = user.id
    mode = MODE_CHOICE.get(user_id)

    if not mode:
        await update.message.reply_text("Сначала выберите режим через /start.")
        return

    document = update.message.document
    file = await context.bot.get_file(document.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        await file.download_to_drive(f.name)
        data_file = f.name

    if mode == "chunk":
        await process_in_parts(update, context, data_file, chunk_size=1000)
    elif mode == "chunk500":
        await process_in_parts(update, context, data_file, chunk_size=500)
    elif mode == "chunk250":
        await process_in_parts(update, context, data_file, chunk_size=250)
    elif mode == "passport":
        await process_passport_macro(update, context, data_file)
    else:
        await update.message.reply_text("Неизвестный режим обработки.")
        return

    # После обработки показываем кнопки заново
    keyboard = [
        [InlineKeyboardButton("▶️ Обработать на части (1000)", callback_data="chunk")],
        [InlineKeyboardButton("▶️ Обработать на части (500)", callback_data="chunk500")],
        [InlineKeyboardButton("▶️ Обработать на части (250)", callback_data="chunk250")],
        [InlineKeyboardButton("📄 Макрос Пасспорт", callback_data="passport")],
        [InlineKeyboardButton("🔄 Замена ПИНФЛ", callback_data="pinfl_replace")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("Выберите режим обработки:", reply_markup=reply_markup)

async def process_in_parts(update, context, data_file, chunk_size=1000):
    logger.info(f"[LOG] Обработка: разбивка на части по {chunk_size} шт.")
    df = pd.read_excel(data_file, header=None, skiprows=3)

    def fix_code(x):
        try:
            s = str(int(float(x)))
            if len(s) == 5:
                return "0" + s
            return x
        except:
            return x

    df[10] = df[10].apply(fix_code)

    seen = set()
    for idx, val in df[0].items():
        val = str(val).strip()
        if val and val in seen:
            df.loc[idx, 0:7] = None
        else:
            seen.add(val)

    parts = [df[i:i + chunk_size] for i in range(0, len(df), chunk_size)]

    output_files = []
    for idx, part in enumerate(parts):
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        for r_idx, row in enumerate(dataframe_to_rows(part, index=False, header=False), start=4):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        start_range = idx * chunk_size
        filename = f"AllPackageEC_{start_range}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)
        output_files.append(output_path)
        logger.info(f"[LOG] Сохранён файл: {filename}")

    zip_path = os.path.join(tempfile.gettempdir(), f"AllPackageEC_{update.message.from_user.username}.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for file_path in output_files:
            zipf.write(file_path, os.path.basename(file_path))

    await update.message.reply_text("Обработка завершена. Архив отправляется...")
    await context.bot.send_document(chat_id=update.message.chat_id, document=open(zip_path, 'rb'))

async def process_passport_macro(update, context, data_file):
    logger.info("[LOG] Выполняется макрос 'Пасспорт'")
    wb = load_workbook(data_file)
    ws = wb.active

    valid_start = "123456789MRTGKZECUVFBNDGHJLKQIP"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        val = str(row[4].value).strip() if row[4].value else ""
        if val and val[0].upper() in valid_start:
            row[4].value = "AB0663236"
            row[5].value = "23,12,1988"

    output_path = os.path.join(tempfile.gettempdir(), f"PassportUpdated_{update.message.from_user.username}.xlsx")
    wb.save(output_path)

    await update.message.reply_text("Макрос выполнен. Файл отправляется...")
    await context.bot.send_document(chat_id=update.message.chat_id, document=open(output_path, 'rb'))

# --- Обработка замены ПИНФЛ ---

async def pinfl_file_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    document = update.message.document
    file = await context.bot.get_file(document.file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as f:
        await file.download_to_drive(f.name)
        data_file = f.name

    step = context.user_data.get('pinfl_step')
    if step == PINFL_SOURCE:
        context.user_data['pinfl_source'] = data_file
        context.user_data['pinfl_step'] = PINFL_PINFL
        await update.message.reply_text("Теперь загрузите файл с результатами ПИНФЛ.")
        return PINFL_STEP
    elif step == PINFL_PINFL:
        context.user_data['pinfl_pinfl'] = data_file
        await update.message.reply_text("Файлы получены. Выполняю замену ПИНФЛ...")
        await replace_pinfl(update, context)
        context.user_data.clear()
        # После обработки показываем кнопки заново
        keyboard = [
            [InlineKeyboardButton("▶️ Обработать на части (1000)", callback_data="chunk")],
            [InlineKeyboardButton("▶️ Обработать на части (500)", callback_data="chunk500")],
            [InlineKeyboardButton("▶️ Обработать на части (250)", callback_data="chunk250")],
            [InlineKeyboardButton("📄 Макрос Пасспорт", callback_data="passport")],
            [InlineKeyboardButton("🔄 Замена ПИНФЛ", callback_data="pinfl_replace")],
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text("Выберите режим обработки:", reply_markup=reply_markup)
        return ConversationHandler.END
    else:
        await update.message.reply_text("Что-то пошло не так. Начните заново с /start.")
        return ConversationHandler.END

async def replace_pinfl(update: Update, context: ContextTypes.DEFAULT_TYPE):
    source_file = context.user_data.get('pinfl_source')
    pinfl_file = context.user_data.get('pinfl_pinfl')
    if not source_file or not pinfl_file:
        await update.message.reply_text("Ошибка: файлы не найдены. Попробуйте заново /start.")
        return

    # Загружаем файл pinfl, без заголовков
    df2 = pd.read_excel(pinfl_file, header=None)

    # Создаём словарь: паспорт -> ПИНФЛ
    passport_to_pinfl = dict(
        zip(df2.iloc[:, 8].astype(str).str.strip().str.upper(), df2.iloc[:, 9])
    )

    # Загружаем оригинальный Excel-файл со стилями
    wb = load_workbook(filename=source_file)
    ws = wb.active  # активный лист

    # Допустимые символы начала паспорта
    valid_start = tuple('0123456789KJTIFHBMNCXZSDQWRYUPLE')

    # Лог замен
    replacements = []

    # Проходим по строкам начиная со второй
    for row in ws.iter_rows(min_row=2):
        cell_e = row[4]  # колонка E (паспорт)
        cell_f = row[5]  # колонка F (дата рождения)
        val = cell_e.value

        # Если ячейка пустая — ставим паспорт и дату
        if val is None or str(val).strip() == '':
            cell_e.value = 'AB0663236'
            if not isinstance(cell_f, MergedCell):
                cell_f.value = '23.12.1988'
            else:
                print(f"⚠️ Пропущена объединённая ячейка в строке {cell_f.row}")
            continue

        # Если значение валидное — пробуем заменить по словарю
        key = str(val).strip().upper()
        if key.startswith(valid_start):
            pinfl = passport_to_pinfl.get(key)
            if pinfl:
                replacements.append((val, pinfl))
                cell_e.value = pinfl

    # Сохраняем результат
    output_file = os.path.join(tempfile.gettempdir(), f"AllPackageEC_GOOD_{update.message.from_user.username}.xlsx")
    wb.save(output_file)

    # Запись лога замен
    log_path = os.path.join(tempfile.gettempdir(), f"замены_log_{update.message.from_user.username}.txt")
    with open(log_path, 'w', encoding='utf-8') as log_file:
        for old, new in replacements:
            log_file.write(f'{old} → {new}
')

    await update.message.reply_text(f'✅ Готово! Файл сохранён. Отправляю... Заменено {len(replacements)} паспортов.')
    await context.bot.send_document(chat_id=update.message.chat_id, document=open(output_file, 'rb'))
    await context.bot.send_document(chat_id=update.message.chat_id, document=open(log_path, 'rb'))

def main():
    app = ApplicationBuilder().token("7872241701:AAF633V3rjyXTJkD8F0lEW13nDtAqHoqeic").build()

    conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(mode_selected)],
        states={
            PINFL_STEP: [MessageHandler(filters.Document.FileExtension("xlsx"), pinfl_file_received)],
        },
        fallbacks=[CommandHandler('start', start)],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(conv_handler)
    app.add_handler(MessageHandler(filters.Document.FileExtension("xlsx"), handle_file))

    logger.info("[LOG] Бот запущен...")
    app.run_polling()

if __name__ == '__main__':
    main()
